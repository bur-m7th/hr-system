#!/usr/bin/env python3
"""
export.py — HR Payroll System Excel Exporter
─────────────────────────────────────────────
Usage:
  python3 export.py --all
  python3 export.py --employee <id or name>
  python3 export.py --month 2026-02
  python3 export.py --department "Engineering"
  python3 export.py --month 2026-02 --department "Engineering"

Options:
  --db PATH         Path to SQLite DB (default: /app/db/hrpayroll.db)
  --out DIR         Output directory   (default: /app/generated/exports)
  --all             Export everything
  --employee VALUE  Employee ID or partial name
  --month YYYY-MM   Filter by month
  --department NAME Filter by department
"""

import argparse
import os
import sqlite3
import sys
from datetime import datetime
from pathlib import Path

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("ERROR: openpyxl not installed. Run: pip3 install openpyxl")
    sys.exit(1)

# ── Styling constants ─────────────────────────────────────────────────────────
HEADER_FILL   = PatternFill("solid", fgColor="1F3864")
HEADER_FONT   = Font(bold=True, color="FFFFFF", name="Calibri", size=11)
ALT_FILL      = PatternFill("solid", fgColor="EBF0FA")
NORMAL_FILL   = PatternFill("solid", fgColor="FFFFFF")
BORDER_SIDE   = Side(style="thin", color="BFBFBF")
CELL_BORDER   = Border(left=BORDER_SIDE, right=BORDER_SIDE,
                       top=BORDER_SIDE, bottom=BORDER_SIDE)
TOTAL_FONT    = Font(bold=True, name="Calibri", size=11)
TOTAL_FILL    = PatternFill("solid", fgColor="D9E1F2")

def styled_header(ws, columns):
    ws.append(columns)
    for cell in ws[1]:
        cell.font    = HEADER_FONT
        cell.fill    = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border  = CELL_BORDER
    ws.row_dimensions[1].height = 30

def style_rows(ws, start_row=2):
    for i, row in enumerate(ws.iter_rows(min_row=start_row, max_row=ws.max_row)):
        fill = ALT_FILL if i % 2 == 0 else NORMAL_FILL
        for cell in row:
            cell.fill   = fill
            cell.border = CELL_BORDER
            cell.alignment = Alignment(vertical="center")

def auto_width(ws):
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                max_len = max(max_len, len(str(cell.value or "")))
            except:
                pass
        ws.column_dimensions[col_letter].width = min(max(max_len + 4, 12), 40)

def add_summary_row(ws, label="TOTAL"):
    ws.append([])  # blank separator
    row_idx = ws.max_row + 1

def get_db(db_path):
    if not Path(db_path).exists():
        print(f"ERROR: Database not found at {db_path}")
        sys.exit(1)
    conn = sqlite3.connect(db_path)
    conn.row_factory = sqlite3.Row
    return conn

def get_tables(conn):
    cur = conn.execute("SELECT name FROM sqlite_master WHERE type='table' AND name NOT LIKE '_migrations'")
    return [r[0] for r in cur.fetchall()]

def get_columns(conn, table):
    cur = conn.execute(f"PRAGMA table_info({table})")
    return [r[1] for r in cur.fetchall()]

# ─────────────────────────────────────────────────────────────────────────────
#  EXPORT FUNCTIONS
# ─────────────────────────────────────────────────────────────────────────────

def export_employees(conn, out_dir, name_filter=None, dept_filter=None):
    """Export employee records"""
    tables = get_tables(conn)

    # Try to find the employees table (might be named differently)
    emp_table = next((t for t in tables if "employee" in t.lower()), None)
    if not emp_table:
        print("WARNING: No employees table found, skipping employee export")
        return None

    cols = get_columns(conn, emp_table)

    # Build query
    conditions = []
    params = []
    if name_filter:
        # Try name columns
        name_col = next((c for c in cols if "name" in c.lower()), None)
        id_col   = next((c for c in cols if c.lower() == "id"), None)
        if name_col:
            conditions.append(f"({name_col} LIKE ? OR {id_col} = ?)")
            params += [f"%{name_filter}%", name_filter]
    if dept_filter:
        dept_col = next((c for c in cols if "dept" in c.lower() or "department" in c.lower()), None)
        if dept_col:
            conditions.append(f"{dept_col} LIKE ?")
            params.append(f"%{dept_filter}%")

    where = f"WHERE {' AND '.join(conditions)}" if conditions else ""
    query = f"SELECT * FROM {emp_table} {where} ORDER BY 1"
    rows  = conn.execute(query, params).fetchall()

    if not rows:
        print("  No employee records matched filter.")
        return None

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Employees"

    # Filter out binary/encrypted columns for display
    display_cols = [c for c in cols if not any(x in c.lower() for x in ["photo", "blob", "key", "hash", "password"])]
    col_indices  = [cols.index(c) for c in display_cols]

    styled_header(ws, display_cols)
    for row in rows:
        ws.append([row[i] for i in col_indices])

    style_rows(ws)
    auto_width(ws)

    label = f"_{name_filter}" if name_filter else (f"_dept_{dept_filter}" if dept_filter else "_all")
    filename = out_dir / f"employees{label}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    wb.save(filename)
    print(f"  ✓ Employees exported → {filename}  ({len(rows)} records)")
    return filename

def export_payroll(conn, out_dir, month_filter=None, dept_filter=None, employee_filter=None):
    """Export payroll / payments data"""
    tables = get_tables(conn)

    pay_table = next((t for t in tables if any(x in t.lower() for x in ["pay", "payroll", "payment", "salary", "slip"])), None)
    if not pay_table:
        print(f"WARNING: No payroll table found. Available tables: {tables}")
        return None

    cols = conn.execute(f"PRAGMA table_info({pay_table})").fetchall()
    col_names = [c[1] for c in cols]

    conditions = []
    params = []

    if month_filter:
        # Find date/month column
        date_col = next((c for c in col_names if any(x in c.lower() for x in ["date", "month", "period", "created"])), None)
        if date_col:
            conditions.append(f"strftime('%Y-%m', {date_col}) = ?")
            params.append(month_filter)

    if dept_filter:
        dept_col = next((c for c in col_names if "dept" in c.lower() or "department" in c.lower()), None)
        if dept_col:
            conditions.append(f"{dept_col} LIKE ?")
            params.append(f"%{dept_filter}%")

    if employee_filter:
        # Try joining with employees or direct column
        emp_col = next((c for c in col_names if "employee" in c.lower() or c.lower() == "emp_id"), None)
        name_col = next((c for c in col_names if "name" in c.lower()), None)
        if emp_col:
            conditions.append(f"({emp_col} LIKE ? OR {emp_col} = ?)")
            params += [f"%{employee_filter}%", employee_filter]
        elif name_col:
            conditions.append(f"{name_col} LIKE ?")
            params.append(f"%{employee_filter}%")

    where = f"WHERE {' AND '.join(conditions)}" if conditions else ""
    query = f"SELECT * FROM {pay_table} {where} ORDER BY 1 DESC"
    rows  = conn.execute(query, params).fetchall()

    if not rows:
        print(f"  No payroll records matched filters.")
        return None

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Payroll"

    display_cols = [c for c in col_names if not any(x in c.lower() for x in ["blob", "key", "hash", "password"])]
    col_indices  = [col_names.index(c) for c in display_cols]

    styled_header(ws, display_cols)
    for row in rows:
        ws.append([row[i] for i in col_indices])

    # Add totals for numeric columns
    numeric_cols = []
    for i, c in enumerate(display_cols, start=1):
        if any(x in c.lower() for x in ["salary", "amount", "pay", "bonus", "deduct", "total", "net"]):
            numeric_cols.append(i)

    if numeric_cols:
        total_row = [""] * len(display_cols)
        total_row[0] = "TOTAL"
        for col_i in numeric_cols:
            col_letter = get_column_letter(col_i)
            last_data  = ws.max_row
            total_row[col_i - 1] = f"=SUM({col_letter}2:{col_letter}{last_data})"
        ws.append(total_row)
        for cell in ws[ws.max_row]:
            cell.font   = TOTAL_FONT
            cell.fill   = TOTAL_FILL
            cell.border = CELL_BORDER

    style_rows(ws)
    auto_width(ws)

    parts = []
    if employee_filter: parts.append(f"emp_{employee_filter}")
    if month_filter:    parts.append(month_filter)
    if dept_filter:     parts.append(f"dept_{dept_filter}")
    label = ("_" + "_".join(parts)) if parts else "_all"
    filename = out_dir / f"payroll{label}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    wb.save(filename)
    print(f"  ✓ Payroll exported → {filename}  ({len(rows)} records)")
    return filename

def export_all(conn, out_dir):
    """Export every table to its own sheet in one workbook"""
    tables = get_tables(conn)
    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # remove default sheet

    for table in tables:
        rows = conn.execute(f"SELECT * FROM {table}").fetchall()
        if not rows:
            continue
        cols = [d[0] for d in conn.execute(f"SELECT * FROM {table} LIMIT 0").description]
        display_cols = [c for c in cols if not any(x in c.lower() for x in ["blob", "key", "hash", "password", "photo"])]
        col_indices  = [cols.index(c) for c in display_cols]

        ws = wb.create_sheet(title=table[:31])
        styled_header(ws, display_cols)
        for row in rows:
            ws.append([row[i] for i in col_indices])
        style_rows(ws)
        auto_width(ws)
        print(f"  ✓ Table '{table}' — {len(rows)} rows")

    filename = out_dir / f"hr_full_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    wb.save(filename)
    print(f"\n  ✓ Full export saved → {filename}")
    return filename

# ─────────────────────────────────────────────────────────────────────────────
#  MAIN
# ─────────────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(description="HR Payroll Excel Exporter")
    parser.add_argument("--db",         default="/app/db/hrpayroll.db", help="SQLite DB path")
    parser.add_argument("--out",        default="/app/generated/exports", help="Output directory")
    parser.add_argument("--all",        action="store_true",  help="Export all data")
    parser.add_argument("--employee",   type=str, default=None, help="Employee ID or name")
    parser.add_argument("--month",      type=str, default=None, help="Month filter YYYY-MM")
    parser.add_argument("--department", type=str, default=None, help="Department name")
    args = parser.parse_args()

    out_dir = Path(args.out)
    out_dir.mkdir(parents=True, exist_ok=True)

    conn = get_db(args.db)
    print(f"\nConnected to: {args.db}")
    print(f"Output dir:   {out_dir}\n")

    if args.all:
        print("── Full export ──────────────────────────────")
        export_all(conn, out_dir)
    else:
        if args.employee:
            print(f"── Employee export: {args.employee} ─────────────")
            export_employees(conn, out_dir, name_filter=args.employee)
            export_payroll(conn, out_dir, employee_filter=args.employee)
        elif args.month or args.department:
            print(f"── Payroll export: month={args.month} dept={args.department} ──")
            export_payroll(conn, out_dir,
                           month_filter=args.month,
                           dept_filter=args.department)
            if args.department:
                export_employees(conn, out_dir, dept_filter=args.department)
        else:
            print("No filter specified. Use --all, --employee, --month, or --department")
            print("Run with --help for usage.")
            conn.close()
            sys.exit(1)

    conn.close()
    print("\nDone.")

if __name__ == "__main__":
    main()